import unittest
from pathlib import Path
from tempfile import NamedTemporaryFile

try:
    from openpyxl import Workbook
except ImportError:  # pragma: no cover - dependency optional in test env
    Workbook = None

try:
    from app import etl
except Exception as exc:  # pragma: no cover
    etl = None
    _import_error = exc
else:
    parse_balance_workbook = etl.parse_balance_workbook


def _build_sample_workbook(tmp_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance 2025"
    ws["A1"] = "BALANCE DE ENERGÍA EN MWh - AÑO 2025"
    ws.append([])
    ws.append([])
    ws.append(["DESCRIPCIÓN", "Ene", "Feb", "Mar", "Acumulado"])
    ws.append(["A emp. Distribuidoras", 1000, 1100, None, 2100])
    ws.append(["A clientes Libres", 500, 600, None, 1100])
    ws.append(["COES", 200, 250, None, 450])
    ws.append(["Consumo propio de centrales", 50, 55, None, 105])
    ws.append(["Pérdidas Sistemas Transmisión", 20, 25, None, 45])

    ws2 = wb.create_sheet("Balance 2024")
    ws2["A1"] = "BALANCE DE ENERGÍA EN MWh - AÑO 2024"
    ws2.append([])
    ws2.append([])
    ws2.append(["DESCRIPCIÓN", "Ene", "Feb", "Mar", "Acumulado"])
    ws2.append(["A emp. Distribuidoras", 900, 950, 1000, 2850])
    ws2.append(["A clientes Libres", 400, 420, 430, 1250])
    ws2.append(["COES", 150, 155, 160, 465])
    ws2.append(["Consumo propio de centrales", 30, 32, 33, 95])
    ws2.append(["Pérdidas Sistemas Transmisión", 15, 14, 16, 45])

    with NamedTemporaryFile(delete=False, suffix=".xlsx", dir=tmp_dir) as tmp:
        wb.save(tmp.name)
        return Path(tmp.name)


class ParseBalanceWorkbookTests(unittest.TestCase):
    @unittest.skipUnless(Workbook and etl, "Dependencias (openpyxl/pandas) no disponibles en el entorno de prueba")
    def test_detects_multiple_years(self):
        tmp_dir = Path(self._get_tmp_dir())
        excel_path = _build_sample_workbook(tmp_dir)
        results = parse_balance_workbook(excel_path)
        years = {r.year for r in results}
        self.assertEqual(years, {2024, 2025})

    @unittest.skipUnless(Workbook and etl, "Dependencias (openpyxl/pandas) no disponibles en el entorno de prueba")
    def test_trims_observed_months(self):
        tmp_dir = Path(self._get_tmp_dir())
        excel_path = _build_sample_workbook(tmp_dir)
        result_2025 = next(r for r in parse_balance_workbook(excel_path) if r.year == 2025)
        self.assertEqual(result_2025.observed_months, ["Ene", "Feb"])
        self.assertEqual(result_2025.energy.regulados[:2], [1000, 1100])
        self.assertEqual(result_2025.energy.libres[:2], [500, 600])
        self.assertEqual(result_2025.energy.coes[:2], [200, 250])

    @unittest.skipUnless(Workbook and etl, "Dependencias (openpyxl/pandas) no disponibles en el entorno de prueba")
    def test_sheet_selection_prefers_latest_versions(self):
        tmp_dir = Path(self._get_tmp_dir())
        wb = Workbook()
        wb.remove(wb.active)
        sheet_names = [
            "2016 (rev3)",
            "2017",
            "2018",
            "2018-R1",
            "2018-R2 (LDS)",
            "2019",
            "2019-R1",
            "2019-R2",
            "2020",
            "2020 R1",
            "2020 R2",
            "2021",
            "2021-R1",
            "2021-R2",
            "2022",
            "2022-R1",
            "2022-R2",
            "2023",
            "2023-R1",
            "2024",
            "2024-R1",
            "2025",
            "2025V1",
            "R",
            "Perfil",
        ]
        for name in sheet_names:
            wb.create_sheet(title=name)

        transformer = etl.BalanceTransformer(wb)
        candidates = transformer._select_candidates()
        self.assertEqual(set(candidates.keys()), set(range(2016, 2026)))
        self.assertEqual(candidates[2018].sheet_name, "2018-R2 (LDS)")
        self.assertEqual(candidates[2019].sheet_name, "2019-R2")
        self.assertEqual(candidates[2020].sheet_name, "2020 R2")
        self.assertEqual(candidates[2022].sheet_name, "2022-R2")
        self.assertEqual(candidates[2025].sheet_name, "2025V1")
        self.assertEqual(candidates[2016].sheet_name, "2016 (rev3)")

    @unittest.skipUnless(Workbook and etl, "Dependencias (openpyxl/pandas) no disponibles en el entorno de prueba")
    def test_spot_computed_as_residual(self):
        tmp_dir = Path(self._get_tmp_dir())
        wb = Workbook()
        ws = wb.active
        ws.title = "2019"
        ws["A1"] = "BALANCE DE ENERGÍA EN MWh - AÑO 2019"
        ws.append([])
        ws.append(["DESCRIPCIÓN", "Enero", "Febrero", "Marzo", "Acumulado"])
        ws.append(["Venta de energía", 90, 92, 95, 277])
        ws.append(["A emp. Distribuidoras", 10, 11, 12, 33])
        ws.append(["A clientes Libres", 20, 21, 22, 63])
        ws.append(["COES", 50, 60, 70, 180])
        ws.append(["Compra de energía", None, None, None, None])
        ws.append(["COES", 0, 0, 0, 0])
        ws.append(["Consumo propio de centrales", 5, 6, 7, 18])
        ws.append(["Pérdidas Sistemas Transmisión", 1, 2, 3, 6])

        with NamedTemporaryFile(delete=False, suffix=".xlsx", dir=tmp_dir) as tmp:
            wb.save(tmp.name)
            excel_path = Path(tmp.name)

        result = next(r for r in etl.parse_balance_workbook(excel_path) if r.year == 2019)
        self.assertEqual(result.energy.coes[:3], [60.0, 60.0, 61.0])
        self.assertEqual(result.energy.regulados[:2], [10.0, 11.0])
        self.assertEqual(result.energy.libres[:2], [20.0, 21.0])
        self.assertTrue(result.energy.perdidas[0] > 0)

    def _get_tmp_dir(self) -> str:
        # Generate a dedicated temp dir inside the test run folder
        tmp_file = NamedTemporaryFile(delete=False)
        tmp_dir = Path(tmp_file.name).parent
        Path(tmp_file.name).unlink(missing_ok=True)
        return str(tmp_dir)


if __name__ == "__main__":
    unittest.main()
