import asyncio
import logging
import threading
import time
from pathlib import Path
from typing import Callable, Dict, Iterable, List

from openpyxl import load_workbook
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer

from .balance_transformer import BalanceParseResult, BalanceTransformer
from .config import DATA_DIR, DEFAULT_SOURCE_ID, WATCH_DEBOUNCE_SECONDS, XLSX_RETRY_COUNT, XLSX_RETRY_DELAY
from .db import record_etl_run, save_balance_rows, save_balance_sales_rows, upsert_source

logger = logging.getLogger(__name__)


def parse_balance_workbook(path: Path, source_id: str = DEFAULT_SOURCE_ID) -> List[BalanceParseResult]:
    workbook = load_workbook(path, data_only=True)
    transformer = BalanceTransformer(workbook, source_id=source_id)
    return transformer.transform()


def process_balance_file(path: Path, dispatch_event: Callable[[dict], None], source_id: str = DEFAULT_SOURCE_ID) -> None:
    logger.info("Procesando archivo %s", path)
    warnings: List[str] = []
    try:
        results = parse_with_retries(path, source_id=source_id)
        run_warnings: List[str] = []
        first_result = results[0]
        for result in results:
            warnings = result.warnings
            save_balance_rows(
                result.year,
                result.source_id,
                result.months,
                result.energy.regulados,
                result.energy.libres,
                result.energy.coes,
                result.energy.servicios_aux,
                result.energy.perdidas,
                observed_months=result.observed_months,
                last_month=result.last_month,
            )
            if result.sales:
                save_balance_sales_rows(
                    result.year,
                    result.source_id,
                    result.months,
                    result.sales.regulados,
                    result.sales.libres,
                    result.sales.coes_spot,
                    result.sales.otros,
                )
            run_warnings.extend([f"{result.year}: {w}" for w in warnings])
            dispatch_event(
                {
                    "type": "DATASET_UPDATED",
                    "dataset_id": "balance",
                    "source_id": result.source_id,
                    "year": result.year,
                    "message": f"Balance actualizado para {result.year}",
                    "warnings": warnings,
                }
            )
        upsert_source(first_result.source_id, "balance", path.name, last_ingested=time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()))
        record_etl_run(first_result.source_id, "balance", "SUCCESS", f"Ingestado {len(results)} hojas", run_warnings)
        if run_warnings:
            logger.warning("ETL finalizó con advertencias: %s", run_warnings)
    except FileNotFoundError:
        msg = f"Archivo {path} no encontrado. Se omite."
        logger.warning(msg)
        warnings.append(msg)
        record_etl_run(source_id, "balance", "WARNING", msg, warnings)
    except Exception as exc:  # noqa: BLE001
        msg = f"Error procesando {path}: {exc}"
        logger.exception(msg)
        warnings.append(msg)
        record_etl_run(source_id, "balance", "ERROR", msg, warnings)
        dispatch_event(
            {
                "type": "ETL_ERROR",
                "dataset_id": "balance",
                "source_id": source_id,
                "message": msg,
                "warnings": warnings,
            }
        )


def parse_with_retries(path: Path, source_id: str = DEFAULT_SOURCE_ID) -> List[BalanceParseResult]:
    last_exc: Exception | None = None
    for attempt in range(1, XLSX_RETRY_COUNT + 1):
        try:
            workbook = load_workbook(path, data_only=True)
            transformer = BalanceTransformer(workbook, source_id=source_id)
            return transformer.transform()
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            if attempt >= XLSX_RETRY_COUNT:
                break
            logger.warning("Intento %s/%s fallido (%s). Reintentando en %.1fs...", attempt, XLSX_RETRY_COUNT, exc, XLSX_RETRY_DELAY)
            time.sleep(XLSX_RETRY_DELAY)
    raise last_exc or RuntimeError("Fallo al parsear Excel.")


class ExcelEventHandler(FileSystemEventHandler):
    def __init__(self, dispatch_event: Callable[[dict], None], source_id: str = DEFAULT_SOURCE_ID):
        super().__init__()
        self._timers: Dict[Path, threading.Timer] = {}
        self.dispatch_event = dispatch_event
        self.source_id = source_id

    def _schedule(self, path: Path) -> None:
        if path.name.startswith("~$") or path.suffix.lower() not in {".xlsx", ".xlsm"}:
            return
        if path in self._timers:
            self._timers[path].cancel()
        timer = threading.Timer(WATCH_DEBOUNCE_SECONDS, self._run_etl, args=[path])
        self._timers[path] = timer
        timer.start()

    def _run_etl(self, path: Path) -> None:
        try:
            process_balance_file(path, self.dispatch_event, source_id=self.source_id)
        finally:
            self._timers.pop(path, None)

    def on_created(self, event):
        if not event.is_directory:
            self._schedule(Path(event.src_path))

    def on_modified(self, event):
        if not event.is_directory:
            self._schedule(Path(event.src_path))


def start_watcher(dispatch_event: Callable[[dict], None], source_id: str = DEFAULT_SOURCE_ID) -> Observer:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    handler = ExcelEventHandler(dispatch_event, source_id=source_id)
    observer = Observer()
    observer.schedule(handler, str(DATA_DIR), recursive=False)
    observer.daemon = True
    observer.start()
    logger.info("Watchdog iniciado en %s", DATA_DIR)
    return observer
